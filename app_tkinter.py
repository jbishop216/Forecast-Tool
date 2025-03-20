import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from sqlalchemy import create_engine, Column, Integer, String, Float, Date, Enum
from sqlalchemy.orm import declarative_base, sessionmaker
import enum
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import csv
import os

# Create database engine
engine = create_engine('sqlite:///forecast_tool.db')
Base = declarative_base()
Session = sessionmaker(bind=engine)

def get_session():
    return Session()

def verify_db_connection():
    """Verify database connection and reset it if necessary"""
    try:
        session = get_session()
        # Try a simple query
        session.query(Settings).first()
        session.close()
        return True
    except Exception as e:
        print(f"Database connection error: {str(e)}")
        try:
            # Try to reset the connection
            global engine, Session
            engine = create_engine('sqlite:///forecast_tool.db')
            Session = sessionmaker(bind=engine)
            # Create tables if they don't exist
            Base.metadata.create_all(engine)
            return True
        except Exception as e:
            print(f"Failed to reset database connection: {str(e)}")
            return False

# Enums
class EmploymentType(enum.Enum):
    FTE = "FTE"
    CONTRACTOR = "CONTRACTOR"

class ChangeType(enum.Enum):
    NEW_HIRE = "New Hire"
    CONVERSION = "Conversion"
    TERMINATION = "Termination"

# Database Models
class Employee(Base):
    __tablename__ = 'employees'
    
    id = Column(Integer, primary_key=True)
    name = Column(String, nullable=False)
    manager_code = Column(String, nullable=False)
    cost_center = Column(String, nullable=False)
    employment_type = Column(String, nullable=False)
    start_date = Column(Date, nullable=False)
    end_date = Column(Date)

class ProjectAllocation(Base):
    __tablename__ = 'project_allocations'
    
    id = Column(Integer, primary_key=True)
    manager_code = Column(String, nullable=False)
    cost_center = Column(String, nullable=False)
    work_code = Column(String, nullable=False)
    year = Column(Integer, nullable=False)
    jan = Column(Float, default=0)
    feb = Column(Float, default=0)
    mar = Column(Float, default=0)
    apr = Column(Float, default=0)
    may = Column(Float, default=0)
    jun = Column(Float, default=0)
    jul = Column(Float, default=0)
    aug = Column(Float, default=0)
    sep = Column(Float, default=0)
    oct = Column(Float, default=0)
    nov = Column(Float, default=0)
    dec = Column(Float, default=0)

class Settings(Base):
    __tablename__ = 'settings'
    
    id = Column(Integer, primary_key=True)
    fte_hours = Column(Float, default=34.5)
    contractor_hours = Column(Float, default=39.0)

class GA01Week(Base):
    __tablename__ = 'ga01_weeks'
    
    id = Column(Integer, primary_key=True)
    year = Column(Integer, nullable=False)
    month = Column(Integer, nullable=False)
    weeks = Column(Float, nullable=False)

class PlannedChange(Base):
    __tablename__ = 'planned_changes'
    
    id = Column(Integer, primary_key=True)
    description = Column(String, nullable=False)
    change_type = Column(String, nullable=False)
    effective_date = Column(Date, nullable=False)
    employee_id = Column(Integer)
    target_type = Column(String)
    name = Column(String)
    team = Column(String)
    manager_code = Column(String)
    cost_center = Column(String)
    employment_type = Column(String)
    status = Column(String, nullable=False)

class Forecast(Base):
    __tablename__ = 'forecasts'
    
    id = Column(Integer, primary_key=True)
    year = Column(Integer, nullable=False)
    cost_center = Column(String, nullable=False)
    manager_code = Column(String, nullable=False)
    work_code = Column(String, nullable=False)
    jan = Column(Float, default=0)
    feb = Column(Float, default=0)
    mar = Column(Float, default=0)
    apr = Column(Float, default=0)
    may = Column(Float, default=0)
    jun = Column(Float, default=0)
    jul = Column(Float, default=0)
    aug = Column(Float, default=0)
    sep = Column(Float, default=0)
    oct = Column(Float, default=0)
    nov = Column(Float, default=0)
    dec = Column(Float, default=0)
    total_hours = Column(Float, default=0)

# Create tables
Base.metadata.create_all(engine)

class EmployeeDialog(tk.Toplevel):
    def __init__(self, parent, employee=None):
        super().__init__(parent)
        self.parent = parent
        self.employee = employee
        self.result = None
        
        self.title("Employee")
        self.geometry("450x400")
        self.resizable(True, True)
        
        # Make dialog modal
        self.transient(parent)
        self.grab_set()
        
        # Create form
        frame = ttk.Frame(self, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)
        
        # Name
        ttk.Label(frame, text="Name:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.name_var = tk.StringVar()
        ttk.Entry(frame, textvariable=self.name_var, width=40).grid(row=0, column=1, sticky=tk.W, pady=5)
        
        # Manager Code
        ttk.Label(frame, text="Manager Code:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.manager_code_var = tk.StringVar()
        ttk.Entry(frame, textvariable=self.manager_code_var, width=20).grid(row=1, column=1, sticky=tk.W, pady=5)
        
        # Cost Center
        ttk.Label(frame, text="Cost Center:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.cost_center_var = tk.StringVar()
        ttk.Entry(frame, textvariable=self.cost_center_var, width=20).grid(row=2, column=1, sticky=tk.W, pady=5)
        
        # Employment Type
        ttk.Label(frame, text="Employment Type:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.employment_type_var = tk.StringVar()
        employment_types = [et.value for et in EmploymentType]
        ttk.Combobox(frame, textvariable=self.employment_type_var, values=employment_types, width=20, state="readonly").grid(row=3, column=1, sticky=tk.W, pady=5)
        
        # Start Date
        ttk.Label(frame, text="Start Date (YYYY-MM-DD):").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.start_date_var = tk.StringVar()
        ttk.Entry(frame, textvariable=self.start_date_var, width=20).grid(row=4, column=1, sticky=tk.W, pady=5)
        
        # End Date
        ttk.Label(frame, text="End Date (YYYY-MM-DD):").grid(row=5, column=0, sticky=tk.W, pady=5)
        self.end_date_var = tk.StringVar()
        ttk.Entry(frame, textvariable=self.end_date_var, width=20).grid(row=5, column=1, sticky=tk.W, pady=5)
        
        # Buttons
        button_frame = ttk.Frame(frame)
        button_frame.grid(row=6, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="OK", command=self.on_ok, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=self.on_cancel, width=10).pack(side=tk.LEFT, padx=5)
        
        # If editing, populate fields
        if employee:
            self.name_var.set(employee.name)
            self.manager_code_var.set(employee.manager_code)
            self.cost_center_var.set(employee.cost_center)
            self.employment_type_var.set(employee.employment_type)
            self.start_date_var.set(employee.start_date.strftime("%Y-%m-%d") if employee.start_date else "")
            self.end_date_var.set(employee.end_date.strftime("%Y-%m-%d") if employee.end_date else "")
        else:
            self.start_date_var.set(datetime.now().strftime("%Y-%m-%d"))
        
        # Center the dialog on the parent window
        self.center_on_parent()
    
    def center_on_parent(self):
        """Center the dialog on its parent window"""
        self.update_idletasks()
        parent_x = self.parent.winfo_rootx()
        parent_y = self.parent.winfo_rooty()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        
        width = self.winfo_width()
        height = self.winfo_height()
        
        x = parent_x + (parent_width - width) // 2
        y = parent_y + (parent_height - height) // 2
        
        self.geometry(f"+{x}+{y}")
    
    def on_ok(self):
        try:
            # Validate required fields
            if not self.name_var.get().strip():
                raise ValueError("Name is required")
            if not self.manager_code_var.get().strip():
                raise ValueError("Manager Code is required")
            if not self.cost_center_var.get().strip():
                raise ValueError("Cost Center is required")
            if not self.employment_type_var.get():
                raise ValueError("Employment Type is required")
            if not self.start_date_var.get().strip():
                raise ValueError("Start Date is required")
            
            # Parse dates
            start_date = datetime.strptime(self.start_date_var.get(), "%Y-%m-%d").date()
            end_date = None
            if self.end_date_var.get().strip():
                end_date = datetime.strptime(self.end_date_var.get(), "%Y-%m-%d").date()
                if end_date <= start_date:
                    raise ValueError("End Date must be after Start Date")
            
            self.result = {
                "name": self.name_var.get().strip(),
                "manager_code": self.manager_code_var.get().strip(),
                "cost_center": self.cost_center_var.get().strip(),
                "employment_type": self.employment_type_var.get(),
                "start_date": start_date,
                "end_date": end_date
            }
            self.destroy()
        except ValueError as e:
            messagebox.showerror("Error", str(e))
    
    def on_cancel(self):
        self.result = None
        self.destroy()

class EmployeeTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent, padding="10")
        self.parent = parent
        self.create_widgets()
        
    def create_widgets(self):
        # Create toolbar
        toolbar = ttk.Frame(self, style="Panel.TFrame")
        toolbar.pack(fill=tk.X, pady=(0, 10), padx=5)
        
        # Add decorative toolbar accent
        accent_canvas = tk.Canvas(toolbar, width=3, height=30, highlightthickness=0)
        accent_canvas.pack(side=tk.LEFT, padx=(0, 10), fill=tk.Y)
        # Draw a gradient accent
        for i in range(30):
            # Gradient from primary to secondary 
            r = int(74 + (52-74) * i/30)
            g = int(125 + (195-125) * i/30)
            b = int(186 + (143-186) * i/30)
            color = f'#{r:02x}{g:02x}{b:02x}'
            accent_canvas.create_line(0, i, 3, i, fill=color)
        
        # Button container for better visual grouping
        button_container = ttk.Frame(toolbar, style="Panel.TFrame")
        button_container.pack(side=tk.LEFT, fill=tk.Y)
        
        ttk.Button(button_container, text="Add Employee", command=self.add_employee).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_container, text="Edit Employee", command=self.edit_employee).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_container, text="Delete Employee", command=self.delete_employee).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_container, text="Refresh", command=self.load_employees).pack(side=tk.LEFT, padx=5)
        
        # Create treeview
        self.tree = ttk.Treeview(self, columns=("ID", "Name", "Manager Code", "Cost Center", "Type", "Start Date", "End Date"), show="headings")
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        # Configure treeview columns
        self.tree.heading("ID", text="ID")
        self.tree.heading("Name", text="Name")
        self.tree.heading("Manager Code", text="Manager Code")
        self.tree.heading("Cost Center", text="Cost Center")
        self.tree.heading("Type", text="Type")
        self.tree.heading("Start Date", text="Start Date")
        self.tree.heading("End Date", text="End Date")
        
        # Configure column widths
        self.tree.column("ID", width=50)
        self.tree.column("Name", width=200)
        self.tree.column("Manager Code", width=100)
        self.tree.column("Cost Center", width=100)
        self.tree.column("Type", width=100)
        self.tree.column("Start Date", width=100)
        self.tree.column("End Date", width=100)
        
        # Load initial data
        self.load_employees()
    
    def load_employees(self):
        """Load employees from database"""
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        try:
            session = get_session()
            employees = session.query(Employee).all()
            
            for emp in employees:
                self.tree.insert("", tk.END, values=(
                    emp.id,
                    emp.name,
                    emp.manager_code,
                    emp.cost_center,
                    emp.employment_type,
                    emp.start_date,
                    emp.end_date
                ))
            
            session.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load employees: {str(e)}")
    
    def add_employee(self):
        """Open dialog to add a new employee"""
        dialog = EmployeeDialog(self)
        dialog.wait_window()
        
        if dialog.result:
            try:
                session = get_session()
                
                # Create new employee
                employee = Employee(
                    name=dialog.result["name"],
                    manager_code=dialog.result["manager_code"],
                    cost_center=dialog.result["cost_center"],
                    employment_type=dialog.result["employment_type"],
                    start_date=dialog.result["start_date"],
                    end_date=dialog.result["end_date"]
                )
                
                session.add(employee)
                session.commit()
                session.close()
                
                # Reload data
                self.load_employees()
                
                # Show success message
                messagebox.showinfo("Success", "Employee added successfully.")
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to add employee: {str(e)}")
    
    def edit_employee(self):
        """Open dialog to edit selected employee"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select an employee to edit.")
            return
        
        # Get employee ID from selection
        emp_id = self.tree.item(selection[0])["values"][0]
        
        try:
            session = get_session()
            
            # Get the employee
            employee = session.query(Employee).filter(Employee.id == emp_id).first()
            if employee:
                # Open dialog with current values
                dialog = EmployeeDialog(self, employee)
                session.close()
                
                dialog.wait_window()
                
                if dialog.result:
                    try:
                        session = get_session()
                        
                        # Update employee
                        employee = session.query(Employee).filter(Employee.id == emp_id).first()
                        employee.name = dialog.result["name"]
                        employee.manager_code = dialog.result["manager_code"]
                        employee.cost_center = dialog.result["cost_center"]
                        employee.employment_type = dialog.result["employment_type"]
                        employee.start_date = dialog.result["start_date"]
                        employee.end_date = dialog.result["end_date"]
                        
                        session.commit()
                        session.close()
                        
                        # Reload data
                        self.load_employees()
                        
                        # Show success message
                        messagebox.showinfo("Success", "Employee updated successfully.")
                        
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed to update employee: {str(e)}")
            else:
                messagebox.showerror("Error", "Employee not found.")
                session.close()
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load employee: {str(e)}")
    
    def delete_employee(self):
        """Delete selected employee"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select an employee to delete.")
            return
        
        # Get employee ID from selection
        emp_id = self.tree.item(selection[0])["values"][0]
        
        # Confirm deletion
        if not messagebox.askyesno("Confirm Deletion", f"Are you sure you want to delete employee with ID {emp_id}?"):
            return
        
        # Delete employee
        try:
            session = get_session()
            
            # Get the employee
            employee = session.query(Employee).filter(Employee.id == emp_id).first()
            if employee:
                session.delete(employee)
                session.commit()
                
                # Reload data
                self.load_employees()
                
                # Show success message
                messagebox.showinfo("Success", f"Employee with ID {emp_id} deleted successfully.")
            else:
                messagebox.showerror("Error", "Employee not found.")
            
            session.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete employee: {str(e)}")

class ProjectAllocationTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        
        # Create toolbar
        toolbar = ttk.Frame(self)
        toolbar.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(toolbar, text="Add Allocation", command=self.add_allocation).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="Edit Allocation", command=self.edit_allocation).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="Delete Allocation", command=self.delete_allocation).pack(side=tk.LEFT, padx=2)
        
        # Create treeview with scrollbar
        self.tree_frame = ttk.Frame(self)
        self.tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.tree = ttk.Treeview(self.tree_frame, columns=(
            "manager_code", "year", "cost_center", "work_code",
            "jan", "feb", "mar", "apr", "may", "jun",
            "jul", "aug", "sep", "oct", "nov", "dec"
        ), show="headings")
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(self.tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        # Configure columns
        self.tree.heading("manager_code", text="Manager")
        self.tree.heading("year", text="Year")
        self.tree.heading("cost_center", text="Cost Center")
        self.tree.heading("work_code", text="Work Code")
        self.tree.heading("jan", text="Jan")
        self.tree.heading("feb", text="Feb")
        self.tree.heading("mar", text="Mar")
        self.tree.heading("apr", text="Apr")
        self.tree.heading("may", text="May")
        self.tree.heading("jun", text="Jun")
        self.tree.heading("jul", text="Jul")
        self.tree.heading("aug", text="Aug")
        self.tree.heading("sep", text="Sep")
        self.tree.heading("oct", text="Oct")
        self.tree.heading("nov", text="Nov")
        self.tree.heading("dec", text="Dec")
        
        # Set column widths
        self.tree.column("manager_code", width=100)
        self.tree.column("year", width=60)
        self.tree.column("cost_center", width=100)
        self.tree.column("work_code", width=100)
        for month in ["jan", "feb", "mar", "apr", "may", "jun",
                     "jul", "aug", "sep", "oct", "nov", "dec"]:
            self.tree.column(month, width=50)
        
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        # Load allocations
        self.load_allocations()
    
    def load_allocations(self):
        """Load project allocations from database"""
        try:
            # Clear existing items
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            session = get_session()
            allocations = session.query(ProjectAllocation).all()
            
            for allocation in allocations:
                self.tree.insert("", tk.END, values=(
                    allocation.manager_code,
                    allocation.year,
                    allocation.cost_center,
                    allocation.work_code,
                    allocation.jan,
                    allocation.feb,
                    allocation.mar,
                    allocation.apr,
                    allocation.may,
                    allocation.jun,
                    allocation.jul,
                    allocation.aug,
                    allocation.sep,
                    allocation.oct,
                    allocation.nov,
                    allocation.dec
                ))
            
            session.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load allocations: {str(e)}")
    
    def add_allocation(self):
        """Add a new project allocation"""
        try:
            dialog = ProjectAllocationDialog(self, None, datetime.now().year)
            self.wait_window(dialog)
            
            if dialog.result:
                session = get_session()
                
                # Create new allocation
                allocation = ProjectAllocation(
                    manager_code=dialog.result["manager_code"],
                    year=dialog.result["year"],
                    cost_center=dialog.result["cost_center"],
                    work_code=dialog.result["work_code"],
                    jan=dialog.result["jan"],
                    feb=dialog.result["feb"],
                    mar=dialog.result["mar"],
                    apr=dialog.result["apr"],
                    may=dialog.result["may"],
                    jun=dialog.result["jun"],
                    jul=dialog.result["jul"],
                    aug=dialog.result["aug"],
                    sep=dialog.result["sep"],
                    oct=dialog.result["oct"],
                    nov=dialog.result["nov"],
                    dec=dialog.result["dec"]
                )
                
                session.add(allocation)
                session.commit()
                session.close()
                
                # Reload allocations
                self.load_allocations()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add allocation: {str(e)}")
    
    def edit_allocation(self):
        """Edit selected project allocation"""
        try:
            selected = self.tree.selection()
            if not selected:
                messagebox.showwarning("Warning", "Please select an allocation to edit.")
                return
            
            # Get selected allocation
            values = self.tree.item(selected[0])["values"]
            
            session = get_session()
            allocation = session.query(ProjectAllocation).filter(
                ProjectAllocation.manager_code == values[0],
                ProjectAllocation.year == values[1],
                ProjectAllocation.cost_center == values[2],
                ProjectAllocation.work_code == values[3]
            ).first()
            
            if not allocation:
                session.close()
                messagebox.showerror("Error", "Selected allocation not found in database.")
                return
            
            # Open dialog with current values
            dialog = ProjectAllocationDialog(self, allocation.manager_code, allocation.year, allocation)
            self.wait_window(dialog)
            
            if dialog.result:
                # Update allocation
                allocation.manager_code = dialog.result["manager_code"]
                allocation.year = dialog.result["year"]
                allocation.cost_center = dialog.result["cost_center"]
                allocation.work_code = dialog.result["work_code"]
                allocation.jan = dialog.result["jan"]
                allocation.feb = dialog.result["feb"]
                allocation.mar = dialog.result["mar"]
                allocation.apr = dialog.result["apr"]
                allocation.may = dialog.result["may"]
                allocation.jun = dialog.result["jun"]
                allocation.jul = dialog.result["jul"]
                allocation.aug = dialog.result["aug"]
                allocation.sep = dialog.result["sep"]
                allocation.oct = dialog.result["oct"]
                allocation.nov = dialog.result["nov"]
                allocation.dec = dialog.result["dec"]
                
                session.commit()
            
            session.close()
            
            # Reload allocations
            self.load_allocations()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to edit allocation: {str(e)}")
    
    def delete_allocation(self):
        """Delete selected project allocation"""
        try:
            selected = self.tree.selection()
            if not selected:
                messagebox.showwarning("Warning", "Please select an allocation to delete.")
                return
            
            if not messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this allocation?"):
                return
            
            # Get selected allocation
            values = self.tree.item(selected[0])["values"]
            
            session = get_session()
            allocation = session.query(ProjectAllocation).filter(
                ProjectAllocation.manager_code == values[0],
                ProjectAllocation.year == values[1],
                ProjectAllocation.cost_center == values[2],
                ProjectAllocation.work_code == values[3]
            ).first()
            
            if allocation:
                session.delete(allocation)
                session.commit()
            
            session.close()
            
            # Reload allocations
            self.load_allocations()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete allocation: {str(e)}")

class ProjectAllocationDialog(tk.Toplevel):
    def __init__(self, parent, manager_code, year, allocation=None):
        super().__init__(parent)
        self.parent = parent
        self.manager_code = manager_code
        self.year = year
        self.allocation = allocation
        self.result = None
        
        # Make dialog modal
        self.transient(parent)
        self.grab_set()
        
        self.title("Project Allocation")
        self.geometry("800x600")
        self.resizable(True, True)
        
        # Create form
        frame = ttk.Frame(self, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)
        
        # Manager and Year info
        info_frame = ttk.LabelFrame(frame, text="Allocation Info", padding="10")
        info_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Manager selection
        ttk.Label(info_frame, text="Manager Code:").pack(side=tk.LEFT, padx=5)
        self.manager_var = tk.StringVar(value=manager_code if manager_code else "")
        self.manager_combo = ttk.Combobox(info_frame, textvariable=self.manager_var, width=20, state="readonly")
        self.manager_combo.pack(side=tk.LEFT, padx=5)
        
        # Year selection
        ttk.Label(info_frame, text="Year:").pack(side=tk.LEFT, padx=5)
        self.year_var = tk.StringVar(value=str(year))
        years = [str(y) for y in range(datetime.now().year - 2, datetime.now().year + 3)]
        self.year_combo = ttk.Combobox(info_frame, textvariable=self.year_var, values=years, width=6, state="readonly")
        self.year_combo.pack(side=tk.LEFT, padx=5)
        
        # Project details
        project_frame = ttk.LabelFrame(frame, text="Project Details", padding="10")
        project_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(project_frame, text="Cost Center:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.cost_center_var = tk.StringVar()
        ttk.Entry(project_frame, textvariable=self.cost_center_var, width=20).grid(row=0, column=1, sticky=tk.W, pady=5)
        
        ttk.Label(project_frame, text="Work Code:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.work_code_var = tk.StringVar()
        ttk.Entry(project_frame, textvariable=self.work_code_var, width=20).grid(row=1, column=1, sticky=tk.W, pady=5)
        
        # Monthly allocations
        allocation_frame = ttk.LabelFrame(frame, text="Monthly Allocations", padding="10")
        allocation_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Create month spinboxes
        self.spinboxes = {}
        months = [
            ("January", "jan"), ("February", "feb"), ("March", "mar"),
            ("April", "apr"), ("May", "may"), ("June", "jun"),
            ("July", "jul"), ("August", "aug"), ("September", "sep"),
            ("October", "oct"), ("November", "nov"), ("December", "dec")
        ]
        
        # Create a frame for the spinboxes
        spinbox_frame = ttk.Frame(allocation_frame)
        spinbox_frame.pack(fill=tk.X, pady=5)
        
        for i, (month_name, month_code) in enumerate(months):
            row = i // 3
            col = i % 3
            
            month_frame = ttk.Frame(spinbox_frame)
            month_frame.grid(row=row, column=col, padx=5, pady=5, sticky=tk.W)
            
            ttk.Label(month_frame, text=f"{month_name}:").pack(side=tk.LEFT)
            spinbox = ttk.Spinbox(month_frame, from_=0, to=100, increment=0.5, width=10)
            spinbox.pack(side=tk.LEFT, padx=5)
            self.spinboxes[month_code] = spinbox
        
        # Add paste button
        paste_frame = ttk.Frame(allocation_frame)
        paste_frame.pack(fill=tk.X, pady=10)
        ttk.Button(paste_frame, text="Paste Monthly Values", command=self.paste_values).pack(side=tk.LEFT, padx=5)
        
        # Buttons
        button_frame = ttk.Frame(frame)
        button_frame.pack(pady=10)
        
        ttk.Button(button_frame, text="OK", command=self.on_ok, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=self.on_cancel, width=10).pack(side=tk.LEFT, padx=5)
        
        # Load manager codes
        session = get_session()
        managers = session.query(Employee.manager_code).distinct().all()
        self.manager_combo['values'] = [m[0] for m in managers]
        session.close()
        
        # If editing, populate fields
        if allocation:
            self.manager_var.set(allocation.manager_code)
            self.year_var.set(str(allocation.year))
            self.cost_center_var.set(allocation.cost_center)
            self.work_code_var.set(allocation.work_code)
            
            # Set monthly values
            self.spinboxes["jan"].set(allocation.jan or 0)
            self.spinboxes["feb"].set(allocation.feb or 0)
            self.spinboxes["mar"].set(allocation.mar or 0)
            self.spinboxes["apr"].set(allocation.apr or 0)
            self.spinboxes["may"].set(allocation.may or 0)
            self.spinboxes["jun"].set(allocation.jun or 0)
            self.spinboxes["jul"].set(allocation.jul or 0)
            self.spinboxes["aug"].set(allocation.aug or 0)
            self.spinboxes["sep"].set(allocation.sep or 0)
            self.spinboxes["oct"].set(allocation.oct or 0)
            self.spinboxes["nov"].set(allocation.nov or 0)
            self.spinboxes["dec"].set(allocation.dec or 0)
        
        # Center the dialog
        self.center_on_parent()
    
    def paste_values(self):
        """Paste values from clipboard into monthly spinboxes"""
        try:
            clipboard = self.clipboard_get()
            values = clipboard.strip().split()
            
            # Convert and validate values
            float_values = []
            for val in values:
                try:
                    # Handle both comma and tab delimiters
                    val = val.replace(',', '').strip()
                    float_val = float(val)
                    if float_val < 0 or float_val > 100:
                        raise ValueError(f"Value {float_val} is out of range (0-100)")
                    float_values.append(float_val)
                except ValueError as e:
                    messagebox.showerror("Error", f"Invalid value: {val}\nPlease ensure all values are numbers between 0 and 100.")
                    return
            
            # Check if we have exactly 12 values
            if len(float_values) != 12:
                messagebox.showerror("Error", f"Expected 12 values, got {len(float_values)}.\nPlease copy exactly 12 monthly values.")
                return
            
            # Set the values in the spinboxes
            months = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 
                     'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
            for month, value in zip(months, float_values):
                self.spinboxes[month].set(f"{value:.1f}")
            
            messagebox.showinfo("Success", "Values pasted successfully!")
            
        except tk.TclError:
            messagebox.showerror("Error", "No data in clipboard. Please copy values first.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to paste values: {str(e)}")
    
    def center_on_parent(self):
        """Center the dialog on its parent window"""
        self.update_idletasks()
        
        # Get parent geometry
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        parent_x = self.parent.winfo_rootx()
        parent_y = self.parent.winfo_rooty()
        
        # Calculate position
        width = self.winfo_width()
        height = self.winfo_height()
        
        x = parent_x + (parent_width - width) // 2
        y = parent_y + (parent_height - height) // 2
        
        # Set position only (preserve size)
        self.geometry(f"+{x}+{y}")
    
    def on_ok(self):
        try:
            # Validate required fields
            if not self.manager_var.get():
                raise ValueError("Manager Code is required")
            if not self.year_var.get():
                raise ValueError("Year is required")
            if not self.cost_center_var.get().strip():
                raise ValueError("Cost Center is required")
            if not self.work_code_var.get().strip():
                raise ValueError("Work Code is required")
            
            # Get monthly values
            monthly_values = {}
            for month, spinbox in self.spinboxes.items():
                try:
                    monthly_values[month] = float(spinbox.get())
                except ValueError:
                    raise ValueError(f"Invalid value for {month}")
            
            self.result = {
                "manager_code": self.manager_var.get(),
                "year": int(self.year_var.get()),
                "cost_center": self.cost_center_var.get().strip(),
                "work_code": self.work_code_var.get().strip(),
                **monthly_values
            }
            self.destroy()
        except ValueError as e:
            messagebox.showerror("Error", str(e))
    
    def on_cancel(self):
        """Cancel dialog"""
        self.result = None
        self.destroy()
    
    def get_allocations(self):
        return self.result

class ForecastVisualization(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.create_widgets()
        
    def create_widgets(self):
        # Control frame for options
        control_frame = ttk.Frame(self)
        control_frame.pack(fill=tk.X, expand=False, pady=5)
        
        # Year selector
        ttk.Label(control_frame, text="Year:").pack(side=tk.LEFT, padx=5)
        self.year_var = tk.StringVar(value=str(datetime.now().year))
        years = [str(year) for year in range(datetime.now().year-2, datetime.now().year+3)]
        self.year_combo = ttk.Combobox(control_frame, textvariable=self.year_var, values=years, width=6)
        self.year_combo.pack(side=tk.LEFT, padx=5)
        
        # Chart type selector
        ttk.Label(control_frame, text="Chart Type:").pack(side=tk.LEFT, padx=5)
        self.chart_type_var = tk.StringVar(value="Monthly Forecast")
        chart_types = ["Monthly Forecast", "Manager Allocation", "Employee Type Distribution"]
        self.chart_type_combo = ttk.Combobox(control_frame, textvariable=self.chart_type_var, 
                                        values=chart_types, width=20, state="readonly")
        self.chart_type_combo.pack(side=tk.LEFT, padx=5)
        
        # Create button
        ttk.Button(control_frame, text="Generate Chart", command=self.generate_chart).pack(side=tk.LEFT, padx=10)
        
        # Frame for the chart
        self.chart_frame = ttk.Frame(self)
        self.chart_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Initial message
        ttk.Label(self.chart_frame, text="Select options and click 'Generate Chart' to create a visualization").pack(
            pady=50)
    
    def generate_chart(self):
        # Clear previous chart
        for widget in self.chart_frame.winfo_children():
            widget.destroy()
        
        try:
            chart_type = self.chart_type_var.get()
            year = int(self.year_var.get())
            
            # Create figure and axes
            fig = Figure(figsize=(10, 6), dpi=100)
            ax = fig.add_subplot(111)
            
            if chart_type == "Monthly Forecast":
                self._generate_monthly_forecast_chart(ax, year)
            elif chart_type == "Manager Allocation":
                self._generate_manager_allocation_chart(ax, year)
            elif chart_type == "Employee Type Distribution":
                self._generate_employee_type_distribution(ax, year)
            
            # Add the plot to the window
            canvas = FigureCanvasTkAgg(fig, master=self.chart_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate chart: {str(e)}")
            # Re-add the initial message
            ttk.Label(self.chart_frame, text=f"Error generating chart: {str(e)}").pack(pady=50)
    
    def _generate_monthly_forecast_chart(self, ax, year):
        session = get_session()
        
        # Get forecast data for the year
        forecasts = session.query(Forecast).filter(Forecast.year == year).all()
        
        # Group by month
        monthly_data = {month: 0 for month in range(1, 13)}
        months = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 
                 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
        
        for forecast in forecasts:
            for i, month in enumerate(months, 1):
                monthly_data[i] += getattr(forecast, month, 0) or 0
        
        # Create the chart
        month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        x = range(len(month_names))
        hours = [monthly_data[i] for i in range(1, 13)]
        
        bars = ax.bar(x, hours, color='skyblue')
        ax.set_xlabel('Month')
        ax.set_ylabel('Hours')
        ax.set_title(f'Monthly Forecast Hours for {year}')
        ax.set_xticks(x)
        ax.set_xticklabels(month_names)
        ax.grid(axis='y', linestyle='--', alpha=0.7)
        
        # Add value labels on top of bars
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{int(height):,}',
                   ha='center', va='bottom')
        
        session.close()
    
    def _generate_manager_allocation_chart(self, ax, year):
        session = get_session()
        
        # Get project allocations for the year
        allocations = session.query(ProjectAllocation).filter(ProjectAllocation.year == year).all()
        
        # Group by manager code
        manager_data = {}
        months = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 
                 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
        
        for allocation in allocations:
            if allocation.manager_code not in manager_data:
                manager_data[allocation.manager_code] = 0
            for month in months:
                manager_data[allocation.manager_code] += getattr(allocation, month, 0) or 0
        
        # Sort by hours (descending)
        sorted_data = dict(sorted(manager_data.items(), key=lambda item: item[1], reverse=True))
        
        managers = list(sorted_data.keys())
        hours = list(sorted_data.values())
        
        # Limit to top 10 if there are many
        if len(managers) > 10:
            managers = managers[:10]
            hours = hours[:10]
            ax.set_title(f'Top 10 Manager Allocations for {year}')
        else:
            ax.set_title(f'Manager Allocations for {year}')
        
        # Create the chart
        bars = ax.barh(managers, hours, color='lightgreen')
        ax.set_xlabel('Hours')
        ax.set_ylabel('Manager Code')
        ax.grid(axis='x', linestyle='--', alpha=0.7)
        
        # Add value labels at the end of bars
        for bar in bars:
            width = bar.get_width()
            ax.text(width, bar.get_y() + bar.get_height()/2.,
                   f'{int(width):,}',
                   ha='left', va='center')
        
        session.close()
    
    def _generate_employee_type_distribution(self, ax, year):
        session = get_session()
        
        # Get all active employees for the year
        start_date = datetime(year, 1, 1).date()
        end_date = datetime(year, 12, 31).date()
        
        employees = session.query(Employee).filter(
            Employee.start_date <= end_date,
            (Employee.end_date.is_(None) | (Employee.end_date >= start_date))
        ).all()
        
        # Count by type
        fte_count = sum(1 for e in employees if e.employment_type == "FTE")
        contractor_count = sum(1 for e in employees if e.employment_type == "CONTRACTOR")
        
        # Create pie chart
        labels = ['FTE', 'Contractor']
        sizes = [fte_count, contractor_count]
        colors = ['#ff9999','#66b3ff']
        explode = (0.1, 0)  # explode the 1st slice (FTE)
        
        if sum(sizes) > 0:  # Only create pie chart if there are employees
            wedges, texts, autotexts = ax.pie(sizes, explode=explode, labels=labels, 
                                            colors=colors, autopct='%1.1f%%', 
                                            shadow=True, startangle=90)
            
            # Make percentage labels easier to read
            plt.setp(autotexts, size=9, weight="bold")
            plt.setp(texts, size=9)
            
            ax.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle
        else:
            ax.text(0.5, 0.5, 'No employees found for selected year',
                   ha='center', va='center')
            ax.axis('off')
        
        ax.set_title(f'Employee Type Distribution ({year})')
        
        session.close()

class GA01WeeksTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent, padding="10")
        self.parent = parent
        self.create_widgets()
        
    def create_widgets(self):
        # Create toolbar
        toolbar = ttk.Frame(self, style="Panel.TFrame")
        toolbar.pack(fill=tk.X, pady=(0, 10), padx=5)
        
        # Add decorative toolbar accent
        accent_canvas = tk.Canvas(toolbar, width=3, height=30, highlightthickness=0)
        accent_canvas.pack(side=tk.LEFT, padx=(0, 10), fill=tk.Y)
        # Draw a gradient accent
        for i in range(30):
            r = int(74 + (52-74) * i/30)
            g = int(125 + (195-125) * i/30)
            b = int(186 + (143-186) * i/30)
            color = f'#{r:02x}{g:02x}{b:02x}'
            accent_canvas.create_line(0, i, 3, i, fill=color)
        
        # Button container for better visual grouping
        button_container = ttk.Frame(toolbar, style="Panel.TFrame")
        button_container.pack(side=tk.LEFT, fill=tk.Y)
        
        # Year selector
        ttk.Label(button_container, text="Year:").pack(side=tk.LEFT, padx=5)
        self.year_var = tk.StringVar(value=str(datetime.now().year))
        years = [str(year) for year in range(datetime.now().year-2, datetime.now().year+3)]
        self.year_combo = ttk.Combobox(button_container, textvariable=self.year_var, values=years, width=6)
        self.year_combo.pack(side=tk.LEFT, padx=5)
        
        # Add year change binding
        self.year_combo.bind('<<ComboboxSelected>>', lambda e: self.load_ga01_weeks())
        
        ttk.Button(button_container, text="Load", command=self.load_ga01_weeks).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_container, text="Save", command=self.save_ga01_weeks).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_container, text="Reset to Default", command=self.reset_to_default).pack(side=tk.LEFT, padx=5)
        
        # Create main frame for month entries
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Month entries
        self.spinboxes = {}
        
        for i, month in enumerate(range(1, 13)):
            month_name = datetime(2000, month, 1).strftime('%B')
            
            # Create frame for each month
            month_frame = ttk.Frame(main_frame)
            month_frame.grid(row=i, column=0, sticky=tk.W, pady=5, padx=5)
            
            # Add month label with fixed width
            ttk.Label(month_frame, text=f"{month_name}:", width=15, anchor="e").pack(side=tk.LEFT, padx=(0, 10))
            
            var = tk.DoubleVar(value=4.0)
            spinbox = ttk.Spinbox(month_frame, from_=0, to=5, increment=0.1, 
                                textvariable=var, width=10, 
                                format="%.1f")
            spinbox.pack(side=tk.LEFT)
            
            # Add "weeks" label
            ttk.Label(month_frame, text="weeks").pack(side=tk.LEFT, padx=(5, 0))
            
            self.spinboxes[month] = var
        
        # Add total weeks display
        total_frame = ttk.Frame(main_frame)
        total_frame.grid(row=13, column=0, sticky=tk.W, pady=(15, 5), padx=5)
        
        ttk.Label(total_frame, text="Total Weeks:", width=15, anchor="e").pack(side=tk.LEFT, padx=(0, 10))
        self.total_weeks_var = tk.StringVar(value="48.0")
        ttk.Label(total_frame, textvariable=self.total_weeks_var).pack(side=tk.LEFT)
        ttk.Label(total_frame, text="weeks").pack(side=tk.LEFT, padx=(5, 0))
        
        # Bind spinbox changes to update total
        for var in self.spinboxes.values():
            var.trace_add("write", self.update_total_weeks)
        
        # Load initial data
        self.load_ga01_weeks()
    
    def update_total_weeks(self, *args):
        """Update the total weeks display"""
        try:
            total = sum(float(var.get()) for var in self.spinboxes.values())
            self.total_weeks_var.set(f"{total:.1f}")
        except:
            self.total_weeks_var.set("Error")
    
    def reset_to_default(self):
        """Reset all months to default value of 4.0"""
        if messagebox.askyesno("Confirm Reset", "Are you sure you want to reset all months to 4.0 weeks?"):
            for var in self.spinboxes.values():
                var.set(4.0)
    
    def load_ga01_weeks(self):
        """Load GA01 weeks from database"""
        try:
            year = int(self.year_var.get())
            session = get_session()
            ga01_weeks = session.query(GA01Week).filter(GA01Week.year == year).all()
            
            # Reset to default values
            for month in range(1, 13):
                self.spinboxes[month].set(4.0)
            
            # Update with database values
            for week in ga01_weeks:
                if week.month in self.spinboxes:
                    self.spinboxes[week.month].set(week.weeks)
            
            session.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load GA01 weeks: {str(e)}")
    
    def save_ga01_weeks(self):
        """Save GA01 weeks to database"""
        try:
            year = int(self.year_var.get())
            
            # Validate values
            for month, var in self.spinboxes.items():
                try:
                    value = float(var.get())
                    if value < 0 or value > 5:
                        raise ValueError(f"Invalid value for {datetime(2000, month, 1).strftime('%B')}: {value}")
                except ValueError as e:
                    messagebox.showerror("Validation Error", str(e))
                    return
            
            if messagebox.askyesno("Confirm Save", f"Are you sure you want to save GA01 weeks for {year}?"):
                session = get_session()
                
                # Delete existing weeks for this year
                session.query(GA01Week).filter(GA01Week.year == year).delete()
                
                # Add new weeks
                for month, var in self.spinboxes.items():
                    ga01_week = GA01Week(
                        year=year,
                        month=month,
                        weeks=float(var.get())
                    )
                    session.add(ga01_week)
                
                session.commit()
                session.close()
                
                messagebox.showinfo("Success", f"GA01 weeks for {year} saved successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save GA01 weeks: {str(e)}")

class PlannedChangesTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent, padding="10")
        self.parent = parent
        self.create_widgets()
        
    def create_widgets(self):
        # Create toolbar
        toolbar = ttk.Frame(self, style="Panel.TFrame")
        toolbar.pack(fill=tk.X, pady=(0, 10), padx=5)
        
        # Add decorative toolbar accent
        accent_canvas = tk.Canvas(toolbar, width=3, height=30, highlightthickness=0)
        accent_canvas.pack(side=tk.LEFT, padx=(0, 10), fill=tk.Y)
        # Draw a gradient accent
        for i in range(30):
            r = int(74 + (52-74) * i/30)
            g = int(125 + (195-125) * i/30)
            b = int(186 + (143-186) * i/30)
            color = f'#{r:02x}{g:02x}{b:02x}'
            accent_canvas.create_line(0, i, 3, i, fill=color)
        
        # Button container for better visual grouping
        button_container = ttk.Frame(toolbar, style="Panel.TFrame")
        button_container.pack(side=tk.LEFT, fill=tk.Y)
        
        ttk.Button(button_container, text="Add Change", command=self.add_change).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_container, text="Edit Change", command=self.edit_change).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_container, text="Delete Change", command=self.delete_change).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_container, text="Apply Change", command=self.apply_change).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_container, text="Refresh", command=self.load_changes).pack(side=tk.LEFT, padx=5)
        
        # Create treeview
        self.tree = ttk.Treeview(self, columns=("ID", "Description", "Type", "Effective Date", "Status"), show="headings")
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        # Configure treeview columns
        self.tree.heading("ID", text="ID")
        self.tree.heading("Description", text="Description")
        self.tree.heading("Type", text="Type")
        self.tree.heading("Effective Date", text="Effective Date")
        self.tree.heading("Status", text="Status")
        
        # Configure column widths
        self.tree.column("ID", width=50)
        self.tree.column("Description", width=300)
        self.tree.column("Type", width=100)
        self.tree.column("Effective Date", width=100)
        self.tree.column("Status", width=100)
        
        # Load initial data
        self.load_changes()
    
    def apply_change(self):
        """Apply the selected planned change"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a planned change to apply.")
            return
        
        # Get change ID from selection
        change_id = self.tree.item(selection[0])["values"][0]
        
        try:
            session = get_session()
            
            # Get the planned change
            change = session.query(PlannedChange).filter(PlannedChange.id == change_id).first()
            if not change:
                messagebox.showerror("Error", "Planned change not found.")
                session.close()
                return
            
            if change.status != "Pending":
                messagebox.showwarning("Warning", "This change has already been applied.")
                session.close()
                return
            
            # Apply the change based on its type
            if change.change_type == "New Hire":
                # Create new employee
                employee = Employee(
                    name=change.name,
                    manager_code=change.manager_code,
                    cost_center=change.cost_center,
                    employment_type=change.employment_type,
                    start_date=change.effective_date
                )
                session.add(employee)
                
            elif change.change_type == "Termination":
                # Find and update employee's end date using employee_id
                employee = session.query(Employee).filter(Employee.id == change.employee_id).first()
                
                if employee:
                    employee.end_date = change.effective_date
                else:
                    messagebox.showwarning("Warning", "Employee not found for termination.")
                    session.close()
                    return
                
            elif change.change_type == "Conversion":
                # Find and update employee's employment type
                employee = session.query(Employee).filter(
                    Employee.manager_code == change.manager_code,
                    Employee.cost_center == change.cost_center,
                    Employee.end_date.is_(None)
                ).first()
                
                if employee:
                    employee.employment_type = change.employment_type
                else:
                    messagebox.showwarning("Warning", "Employee not found for conversion.")
                    session.close()
                    return
            
            # Update change status
            change.status = "Applied"
            
            session.commit()
            session.close()
            
            # Refresh the display
            self.load_changes()
            
            # Update employee list and forecast
            if isinstance(self.parent, ForecastApp):
                self.parent.employee_tab.load_employees()
                self.parent.forecast_tab.calculate_forecast()
            
            # Show success message
            messagebox.showinfo("Success", "Planned change applied successfully.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to apply planned change: {str(e)}")
            if 'session' in locals():
                session.close()
    
    def load_changes(self):
        """Load planned changes from database"""
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        try:
            session = get_session()
            changes = session.query(PlannedChange).order_by(PlannedChange.effective_date).all()
            
            for change in changes:
                self.tree.insert("", tk.END, values=(
                    change.id,
                    change.description,
                    change.change_type,
                    change.effective_date.strftime("%Y-%m-%d") if change.effective_date else "",
                    change.status
                ))
            
            session.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load planned changes: {str(e)}")
    
    def add_change(self):
        """Open dialog to add a new planned change"""
        print("Opening planned change dialog...")  # Debug log
        dialog = PlannedChangeDialog(self)
        dialog.wait_window()
        
        if dialog.result:
            try:
                print(f"Dialog returned result: {dialog.result}")  # Debug log
                session = get_session()
                
                try:
                    # Create new planned change
                    change = PlannedChange(
                        description=dialog.result["description"],
                        change_type=dialog.result["change_type"],
                        effective_date=dialog.result["effective_date"],
                        employee_id=dialog.result.get("employee_id"),
                        name=dialog.result.get("name"),
                        team=dialog.result.get("team"),
                        manager_code=dialog.result.get("manager_code"),
                        cost_center=dialog.result.get("cost_center"),
                        employment_type=dialog.result.get("employment_type"),
                        status="Pending"
                    )
                    
                    print(f"Adding change to database: {change.description}, {change.change_type}")  # Debug log
                    session.add(change)
                    session.commit()
                    
                    # Reload data
                    self.load_changes()
                    
                    # Show success message
                    messagebox.showinfo("Success", "Planned change added successfully.")
                    
                except Exception as e:
                    print(f"Error adding planned change: {str(e)}")  # Debug log
                    session.rollback()
                    messagebox.showerror("Error", f"Failed to add planned change: {str(e)}")
                finally:
                    session.close()
                    
            except Exception as e:
                print(f"Error in database session: {str(e)}")  # Debug log
                messagebox.showerror("Error", f"Database error: {str(e)}")
        else:
            # Dialog was cancelled or closed without result
            print("Dialog closed without result")  # Debug log
    
    def edit_change(self):
        """Open dialog to edit selected planned change"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a planned change to edit.")
            return
        
        # Get change ID from selection
        change_id = self.tree.item(selection[0])["values"][0]
        
        try:
            session = get_session()
            
            # Get the planned change
            change = session.query(PlannedChange).filter(PlannedChange.id == change_id).first()
            if change:
                # Open dialog with current values
                dialog = PlannedChangeDialog(self, change)
                session.close()
                
                dialog.wait_window()  # Use dialog.wait_window() for consistency
                
                if dialog.result:
                    try:
                        session = get_session()
                        
                        # Update planned change
                        change = session.query(PlannedChange).filter(PlannedChange.id == change_id).first()
                        change.description = dialog.result["description"]
                        change.change_type = dialog.result["change_type"]
                        change.effective_date = dialog.result["effective_date"]
                        change.employee_id = dialog.result.get("employee_id")
                        change.name = dialog.result.get("name")
                        change.team = dialog.result.get("team")
                        change.manager_code = dialog.result.get("manager_code")
                        change.cost_center = dialog.result.get("cost_center")
                        change.employment_type = dialog.result.get("employment_type")
                        
                        session.commit()
                        session.close()
                        
                        # Reload data
                        self.load_changes()
                        
                        # Show success message
                        messagebox.showinfo("Success", "Planned change updated successfully.")
                        
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed to update planned change: {str(e)}")
                        if 'session' in locals():
                            session.close()
            else:
                messagebox.showerror("Error", "Planned change not found.")
                session.close()
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load planned change: {str(e)}")
            if 'session' in locals():
                session.close()
    
    def delete_change(self):
        """Delete selected planned change"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a planned change to delete.")
            return
        
        # Get change ID from selection
        change_id = self.tree.item(selection[0])["values"][0]
        
        # Confirm deletion
        if not messagebox.askyesno("Confirm Deletion", f"Are you sure you want to delete planned change with ID {change_id}?"):
            return
        
        try:
            session = get_session()
            
            # Get the planned change
            change = session.query(PlannedChange).filter(PlannedChange.id == change_id).first()
            if change:
                session.delete(change)
                session.commit()
                
                # Reload data
                self.load_changes()
                
                # Show success message
                messagebox.showinfo("Success", f"Planned change with ID {change_id} deleted successfully.")
            else:
                messagebox.showerror("Error", "Planned change not found.")
            
            session.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete planned change: {str(e)}")

class ForecastTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent, padding="10")
        self.parent = parent
        self.create_widgets()
        
    def create_widgets(self):
        # Create toolbar
        toolbar = ttk.Frame(self, style="Panel.TFrame")
        toolbar.pack(fill=tk.X, pady=(0, 10), padx=5)
        
        # Add decorative toolbar accent
        accent_canvas = tk.Canvas(toolbar, width=3, height=30, highlightthickness=0)
        accent_canvas.pack(side=tk.LEFT, padx=(0, 10), fill=tk.Y)
        # Draw a gradient accent
        for i in range(30):
            r = int(74 + (52-74) * i/30)
            g = int(125 + (195-125) * i/30)
            b = int(186 + (143-186) * i/30)
            color = f'#{r:02x}{g:02x}{b:02x}'
            accent_canvas.create_line(0, i, 3, i, fill=color)
        
        # Button container for better visual grouping
        button_container = ttk.Frame(toolbar, style="Panel.TFrame")
        button_container.pack(side=tk.LEFT, fill=tk.Y)
        
        # Year selector
        ttk.Label(button_container, text="Year:").pack(side=tk.LEFT, padx=5)
        self.year_var = tk.StringVar(value=str(datetime.now().year))
        years = [str(year) for year in range(datetime.now().year-2, datetime.now().year+3)]
        self.year_combo = ttk.Combobox(button_container, textvariable=self.year_var, values=years, width=6)
        self.year_combo.pack(side=tk.LEFT, padx=5)
        
        # Work code selector
        ttk.Label(button_container, text="Work Code:").pack(side=tk.LEFT, padx=5)
        self.work_code_var = tk.StringVar(value="All")
        self.work_code_combo = ttk.Combobox(button_container, textvariable=self.work_code_var, width=20)
        self.work_code_combo.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_container, text="Calculate Forecast", command=self.calculate_forecast).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_container, text="Refresh", command=self.load_forecast).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_container, text="Export to Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=5)
        
        # Create treeview
        self.tree = ttk.Treeview(self, columns=("Cost Center", "Manager", "Work Code", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Total"), show="headings")
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        # Configure treeview columns
        self.tree.heading("Cost Center", text="Cost Center")
        self.tree.heading("Manager", text="Manager")
        self.tree.heading("Work Code", text="Work Code")
        self.tree.heading("Jan", text="Jan")
        self.tree.heading("Feb", text="Feb")
        self.tree.heading("Mar", text="Mar")
        self.tree.heading("Apr", text="Apr")
        self.tree.heading("May", text="May")
        self.tree.heading("Jun", text="Jun")
        self.tree.heading("Jul", text="Jul")
        self.tree.heading("Aug", text="Aug")
        self.tree.heading("Sep", text="Sep")
        self.tree.heading("Oct", text="Oct")
        self.tree.heading("Nov", text="Nov")
        self.tree.heading("Dec", text="Dec")
        self.tree.heading("Total", text="Total")
        
        # Configure column widths
        self.tree.column("Cost Center", width=100)
        self.tree.column("Manager", width=100)
        self.tree.column("Work Code", width=100)
        self.tree.column("Jan", width=60)
        self.tree.column("Feb", width=60)
        self.tree.column("Mar", width=60)
        self.tree.column("Apr", width=60)
        self.tree.column("May", width=60)
        self.tree.column("Jun", width=60)
        self.tree.column("Jul", width=60)
        self.tree.column("Aug", width=60)
        self.tree.column("Sep", width=60)
        self.tree.column("Oct", width=60)
        self.tree.column("Nov", width=60)
        self.tree.column("Dec", width=60)
        self.tree.column("Total", width=80)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        # Load initial data
        self.load_work_codes()
        self.load_forecast()
    
    def load_work_codes(self):
        """Load work codes from project allocations"""
        try:
            session = get_session()
            work_codes = session.query(ProjectAllocation.work_code).distinct().all()
            work_codes = ["All"] + [wc[0] for wc in work_codes]
            self.work_code_combo["values"] = work_codes
            session.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load work codes: {str(e)}")
    
    def calculate_forecast(self):
        """Calculate forecast based on project allocations, GA01 weeks, and planned changes"""
        try:
            year = int(self.year_var.get())
            work_code = self.work_code_var.get()
            
            session = get_session()
            
            # Get or create settings with default values
            settings = session.query(Settings).first()
            if not settings:
                settings = Settings(fte_hours=34.5, contractor_hours=39.0)
                session.add(settings)
                session.commit()
            
            # Get GA01 weeks for the year
            ga01_weeks = session.query(GA01Week).filter(GA01Week.year == year).all()
            ga01_dict = {week.month: week.weeks for week in ga01_weeks}
            
            # Get project allocations
            query = session.query(ProjectAllocation).filter(ProjectAllocation.year == year)
            if work_code != "All":
                query = query.filter(ProjectAllocation.work_code == work_code)
            allocations = query.all()
            
            if not allocations:
                messagebox.showwarning("No Data", "No project allocations found for the selected criteria.")
                session.close()
                return
            
            # Delete existing forecasts for this year and work code
            if work_code == "All":
                session.query(Forecast).filter(Forecast.year == year).delete()
            else:
                session.query(Forecast).filter(
                    Forecast.year == year,
                    Forecast.work_code == work_code
                ).delete()
            
            # Calculate forecast
            months = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 
                     'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
            
            # Get planned changes for the year
            planned_changes = session.query(PlannedChange).filter(
                PlannedChange.effective_date.between(
                    datetime(year, 1, 1).date(),
                    datetime(year, 12, 31).date()
                ),
                PlannedChange.status == "Pending"
            ).order_by(PlannedChange.effective_date).all()
            
            # Create a dictionary to track active employees and their hours per week
            active_employees = {}  # key: (manager_code, cost_center), value: list of (employee_id, start_date, end_date, hours_per_week)
            
            # Initialize with existing employees
            existing_employees = session.query(Employee).filter(
                Employee.start_date <= datetime(year, 12, 31).date(),
                (Employee.end_date.is_(None) | (Employee.end_date >= datetime(year, 1, 1).date()))
            ).all()
            
            for emp in existing_employees:
                key = (emp.manager_code, emp.cost_center)
                hours = settings.contractor_hours if emp.employment_type == "CONTRACTOR" else settings.fte_hours
                if key not in active_employees:
                    active_employees[key] = []
                active_employees[key].append((emp.id, emp.start_date, emp.end_date, hours))
            
            # Add planned changes to active_employees
            for change in planned_changes:
                key = (change.manager_code, change.cost_center)
                if change.change_type == "New Hire":
                    hours = settings.contractor_hours if change.employment_type == "CONTRACTOR" else settings.fte_hours
                    if key not in active_employees:
                        active_employees[key] = []
                    # New hires don't have an employee_id yet
                    active_employees[key].append((None, change.effective_date, None, hours))
                elif change.change_type == "Termination":
                    if key in active_employees and change.employee_id:
                        # Update the end date of the specific employee being terminated
                        for i, (emp_id, start_date, end_date, hours) in enumerate(active_employees[key]):
                            if emp_id == change.employee_id:
                                active_employees[key][i] = (emp_id, start_date, change.effective_date, hours)
                                break
                elif change.change_type == "Conversion":
                    if key in active_employees and change.employee_id:
                        # Add new entry with updated hours for the specific employee
                        for i, (emp_id, start_date, end_date, hours) in enumerate(active_employees[key]):
                            if emp_id == change.employee_id:
                                hours = settings.contractor_hours if change.employment_type == "CONTRACTOR" else settings.fte_hours
                                active_employees[key][i] = (emp_id, start_date, change.effective_date, hours)
                                active_employees[key].append((emp_id, change.effective_date, None, hours))
                                break
            
            for allocation in allocations:
                total_hours = 0
                forecast = Forecast(
                    year=year,
                    cost_center=allocation.cost_center,
                    manager_code=allocation.manager_code,
                    work_code=allocation.work_code
                )
                
                key = (allocation.manager_code, allocation.cost_center)
                
                # Calculate monthly hours based on GA01 weeks
                for i, month in enumerate(months, 1):
                    # Get GA01 weeks for the month (default to 4.0 if not set)
                    weeks = float(ga01_dict.get(i, 4.0))
                    
                    # Calculate month start and end dates
                    month_start = datetime(year, i, 1).date()
                    month_end = datetime(year, i + 1, 1).date() if i < 12 else datetime(year + 1, 1, 1).date()
                    
                    # Calculate total work hours for the month based on active employees
                    total_work_hours = 0
                    if key in active_employees:
                        for emp_id, start_date, end_date, hours_per_week in active_employees[key]:
                            # For each month, if the employee is active during any part of the month,
                            # count them for the full month
                            if (start_date <= month_end and 
                                (end_date is None or end_date >= month_start)):
                                total_work_hours += hours_per_week * weeks
                    
                    # Get allocation hours to subtract
                    allocation_hours = float(getattr(allocation, month, 0) or 0)
                    
                    # Calculate forecast hours by subtracting allocation hours
                    forecast_hours = max(0, total_work_hours - allocation_hours)
                    
                    # Set forecast hours for the month
                    setattr(forecast, month, round(forecast_hours, 2))
                    total_hours += forecast_hours
                
                forecast.total_hours = round(total_hours, 2)
                session.add(forecast)
            
            session.commit()
            session.close()
            
            # Reload data
            self.load_forecast()
            
            # Show success message
            messagebox.showinfo("Success", "Forecast calculated successfully.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to calculate forecast: {str(e)}")
            if 'session' in locals():
                session.close()
    
    def load_forecast(self):
        """Load forecast data from database"""
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        try:
            year = int(self.year_var.get())
            work_code = self.work_code_var.get()
            
            session = get_session()
            
            # Get forecasts
            query = session.query(Forecast).filter(Forecast.year == year)
            if work_code != "All":
                query = query.filter(Forecast.work_code == work_code)
            forecasts = query.all()
            
            for forecast in forecasts:
                self.tree.insert("", tk.END, values=(
                    forecast.cost_center,
                    forecast.manager_code,
                    forecast.work_code,
                    round(forecast.jan, 2) if forecast.jan else 0,
                    round(forecast.feb, 2) if forecast.feb else 0,
                    round(forecast.mar, 2) if forecast.mar else 0,
                    round(forecast.apr, 2) if forecast.apr else 0,
                    round(forecast.may, 2) if forecast.may else 0,
                    round(forecast.jun, 2) if forecast.jun else 0,
                    round(forecast.jul, 2) if forecast.jul else 0,
                    round(forecast.aug, 2) if forecast.aug else 0,
                    round(forecast.sep, 2) if forecast.sep else 0,
                    round(forecast.oct, 2) if forecast.oct else 0,
                    round(forecast.nov, 2) if forecast.nov else 0,
                    round(forecast.dec, 2) if forecast.dec else 0,
                    round(forecast.total_hours, 2) if forecast.total_hours else 0
                ))
            
            session.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load forecast: {str(e)}")
    
    def export_to_excel(self):
        """Export forecast data to Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            # Create workbook and select active sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Forecast"
            
            # Write headers
            headers = ["Cost Center", "Manager", "Work Code", "Jan", "Feb", "Mar", "Apr", "May", "Jun", 
                      "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Total"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Write data
            row = 2
            for item in self.tree.get_children():
                values = self.tree.item(item)["values"]
                for col, value in enumerate(values, 1):
                    ws.cell(row=row, column=col, value=value)
                row += 1
            
            # Save file
            from tkinter import filedialog
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Export Forecast"
            )
            
            if filename:
                wb.save(filename)
                messagebox.showinfo("Success", f"Forecast exported to {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export forecast: {str(e)}")

class PlannedChangeDialog(tk.Toplevel):
    def __init__(self, parent, change=None):
        super().__init__(parent)
        self.parent = parent
        self.change = change
        self.result = None
        self.employee_list = []  # Initialize employee_list to empty list
        
        self.title("Planned Change")
        self.geometry("500x600")
        self.resizable(True, True)
        
        # Make dialog modal
        self.transient(parent)
        self.grab_set()
        
        # Create form
        frame = ttk.Frame(self, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)
        
        # Description
        ttk.Label(frame, text="Description:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.description_var = tk.StringVar()
        ttk.Entry(frame, textvariable=self.description_var, width=40).grid(row=0, column=1, sticky=tk.W, pady=5)
        
        # Change Type
        ttk.Label(frame, text="Change Type:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.change_type_var = tk.StringVar()
        change_types = [ct.value for ct in ChangeType]
        self.change_type_combo = ttk.Combobox(frame, textvariable=self.change_type_var, values=change_types, width=20, state="readonly")
        self.change_type_combo.grid(row=1, column=1, sticky=tk.W, pady=5)
        
        # Bind change type selection to update form
        self.change_type_combo.bind('<<ComboboxSelected>>', self.on_change_type_selected)
        
        # Effective Date
        ttk.Label(frame, text="Effective Date (YYYY-MM-DD):").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.effective_date_var = tk.StringVar()
        ttk.Entry(frame, textvariable=self.effective_date_var, width=20).grid(row=2, column=1, sticky=tk.W, pady=5)
        
        # Employee Details Frame
        self.employee_frame = ttk.LabelFrame(frame, text="Employee Details", padding="10")
        self.employee_frame.grid(row=3, column=0, columnspan=2, sticky=tk.EW, pady=10)
        
        # Employee Selector (for Termination)
        self.employee_selector_frame = ttk.Frame(self.employee_frame)
        self.employee_selector_frame.grid(row=0, column=0, columnspan=2, sticky=tk.EW, pady=5)
        ttk.Label(self.employee_selector_frame, text="Select Employee:").grid(row=0, column=0, sticky=tk.W)
        self.employee_var = tk.StringVar()
        self.employee_combo = ttk.Combobox(self.employee_selector_frame, textvariable=self.employee_var, width=40, state="readonly")
        self.employee_combo.grid(row=0, column=1, sticky=tk.W, padx=5)
        
        # Name
        ttk.Label(self.employee_frame, text="Name:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.name_var = tk.StringVar()
        self.name_entry = ttk.Entry(self.employee_frame, textvariable=self.name_var, width=40)
        self.name_entry.grid(row=1, column=1, sticky=tk.W, pady=5)
        
        # Team
        ttk.Label(self.employee_frame, text="Team:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.team_var = tk.StringVar()
        self.team_entry = ttk.Entry(self.employee_frame, textvariable=self.team_var, width=20)
        self.team_entry.grid(row=2, column=1, sticky=tk.W, pady=5)
        
        # Manager Code
        ttk.Label(self.employee_frame, text="Manager Code:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.manager_code_var = tk.StringVar()
        self.manager_code_entry = ttk.Entry(self.employee_frame, textvariable=self.manager_code_var, width=20)
        self.manager_code_entry.grid(row=3, column=1, sticky=tk.W, pady=5)
        
        # Cost Center
        ttk.Label(self.employee_frame, text="Cost Center:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.cost_center_var = tk.StringVar()
        self.cost_center_entry = ttk.Entry(self.employee_frame, textvariable=self.cost_center_var, width=20)
        self.cost_center_entry.grid(row=4, column=1, sticky=tk.W, pady=5)
        
        # Employment Type
        ttk.Label(self.employee_frame, text="Employment Type:").grid(row=5, column=0, sticky=tk.W, pady=5)
        self.employment_type_var = tk.StringVar()
        employment_types = [et.value for et in EmploymentType]
        self.employment_type_combo = ttk.Combobox(self.employee_frame, textvariable=self.employment_type_var, values=employment_types, width=20, state="readonly")
        self.employment_type_combo.grid(row=5, column=1, sticky=tk.W, pady=5)
        
        # Buttons
        button_frame = ttk.Frame(frame)
        button_frame.grid(row=4, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="OK", command=self.on_ok, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=self.on_cancel, width=10).pack(side=tk.LEFT, padx=5)
        
        # If editing, populate fields
        if change:
            self.description_var.set(change.description)
            self.change_type_var.set(change.change_type)
            self.effective_date_var.set(change.effective_date.strftime("%Y-%m-%d") if change.effective_date else "")
            self.name_var.set(change.name or "")
            self.team_var.set(change.team or "")
            self.manager_code_var.set(change.manager_code or "")
            self.cost_center_var.set(change.cost_center or "")
            self.employment_type_var.set(change.employment_type or "")
            
            # If it's a termination, select the employee
            if change.change_type == "Termination" and change.employee_id:
                self.load_active_employees()
                for display_name, employee in self.employee_list:
                    if employee.id == change.employee_id:
                        self.employee_var.set(display_name)
                        break
        else:
            self.effective_date_var.set(datetime.now().strftime("%Y-%m-%d"))
        
        # Load active employees for termination
        self.load_active_employees()
        
        # Center the dialog on the parent window
        self.center_on_parent()
    
    def load_active_employees(self):
        """Load active employees for the termination dropdown"""
        try:
            session = get_session()
            employees = session.query(Employee).filter(
                Employee.end_date.is_(None)
            ).all()
            
            # Format employee names for dropdown
            employee_list = []
            for emp in employees:
                display_name = f"{emp.name} ({emp.manager_code} - {emp.cost_center})"
                employee_list.append((display_name, emp))
            
            self.employee_list = employee_list
            self.employee_combo['values'] = [name for name, _ in employee_list]
            
            session.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load employees: {str(e)}")
    
    def on_change_type_selected(self, event=None):
        """Handle change type selection"""
        change_type = self.change_type_var.get()
        
        if change_type == "Termination":
            # Show employee selector, hide other fields
            self.employee_selector_frame.grid()
            
            # Hide other employee fields
            for widget in [self.name_entry, self.team_entry, self.manager_code_entry, 
                         self.cost_center_entry, self.employment_type_combo]:
                widget.grid_remove()
            
            # Bind employee selection
            self.employee_combo.bind('<<ComboboxSelected>>', self.on_employee_selected)
            
            print("Termination selected, showing employee dropdown")  # Debug log
        else:
            # Hide employee selector, show other fields
            self.employee_selector_frame.grid_remove()
            
            # Show other employee fields
            for widget in [self.name_entry, self.team_entry, self.manager_code_entry, 
                         self.cost_center_entry, self.employment_type_combo]:
                widget.grid()
            
            # Unbind employee selection
            self.employee_combo.unbind('<<ComboboxSelected>>')
            
            print(f"{change_type} selected, showing employee detail fields")  # Debug log
    
    def on_employee_selected(self, event=None):
        """Handle employee selection"""
        selection = self.employee_var.get()
        for display_name, employee in self.employee_list:
            if display_name == selection:
                self.name_var.set(employee.name)
                self.manager_code_var.set(employee.manager_code)
                self.cost_center_var.set(employee.cost_center)
                self.employment_type_var.set(employee.employment_type)
                break
    
    def on_ok(self):
        try:
            print("OK button clicked in dialog")  # Debug log
            
            # Validate required fields
            if not self.description_var.get().strip():
                raise ValueError("Description is required")
            if not self.change_type_var.get():
                raise ValueError("Change Type is required")
            if not self.effective_date_var.get().strip():
                raise ValueError("Effective Date is required")
            
            # Parse date
            try:
                effective_date = datetime.strptime(self.effective_date_var.get(), "%Y-%m-%d").date()
            except ValueError:
                raise ValueError("Invalid date format. Please use YYYY-MM-DD")
            
            # Get selected employee for termination
            employee_id = None
            if self.change_type_var.get() == "Termination":
                if not self.employee_var.get():
                    raise ValueError("Please select an employee to terminate")
                selection = self.employee_var.get()
                found = False
                for display_name, employee in self.employee_list:
                    if display_name == selection:
                        employee_id = employee.id
                        found = True
                        break
                if not found:
                    raise ValueError("Selected employee not found")
                
                print(f"Termination change for employee ID: {employee_id}")  # Debug log
            
            # For New Hire and Conversion, validate required fields
            if self.change_type_var.get() in ["New Hire", "Conversion"]:
                if not self.name_var.get().strip():
                    raise ValueError("Name is required")
                if not self.manager_code_var.get().strip():
                    raise ValueError("Manager Code is required")
                if not self.cost_center_var.get().strip():
                    raise ValueError("Cost Center is required")
                if not self.employment_type_var.get():
                    raise ValueError("Employment Type is required")
                
                print(f"{self.change_type_var.get()} change for: {self.name_var.get()}")  # Debug log
            
            self.result = {
                "description": self.description_var.get().strip(),
                "change_type": self.change_type_var.get(),
                "effective_date": effective_date,
                "employee_id": employee_id,
                "name": self.name_var.get().strip() or None,
                "team": self.team_var.get().strip() or None,
                "manager_code": self.manager_code_var.get().strip() or None,
                "cost_center": self.cost_center_var.get().strip() or None,
                "employment_type": self.employment_type_var.get() or None,
                "status": "Pending"  # Add status field
            }
            print(f"Dialog result: {self.result}")  # Debug log
            self.destroy()
        except ValueError as e:
            messagebox.showerror("Error", str(e))
            print(f"Validation error: {str(e)}")  # Debug log
        except Exception as e:
            print(f"Unexpected error in on_ok: {str(e)}")  # Debug log
            messagebox.showerror("Error", str(e))
    
    def on_cancel(self):
        """Cancel dialog"""
        self.result = None
        self.destroy()
        
    def center_on_parent(self):
        """Center the dialog on the parent window"""
        self.update_idletasks()
        
        # Get parent geometry
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        parent_x = self.parent.winfo_rootx()
        parent_y = self.parent.winfo_rooty()
        
        # Calculate position
        width = self.winfo_width()
        height = self.winfo_height()
        
        x = parent_x + (parent_width - width) // 2
        y = parent_y + (parent_height - height) // 2
        
        # Set position only (preserve size)
        self.geometry(f"+{x}+{y}")

class ForecastApp(tk.Tk):
    def __init__(self):
        super().__init__()
        
        self.title("Forecast Tool")
        self.geometry("1200x800")
        
        # Verify database connection
        if not verify_db_connection():
            messagebox.showerror("Database Error", "Failed to connect to database. The application may not function correctly.")
        
        # Initialize settings if needed
        self.init_settings()
        
        # Configure modern style
        style = ttk.Style()
        style.configure(".", font=("Helvetica", 10))
        style.configure("Treeview", rowheight=25)
        style.configure("Treeview.Heading", font=("Helvetica", 10, "bold"))
        style.configure("TButton", padding=6)
        style.configure("TEntry", padding=6)
        style.configure("TCombobox", padding=6)
        style.configure("Panel.TFrame", background="#f0f0f0")
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create tabs
        self.employee_tab = EmployeeTab(self.notebook)
        self.notebook.add(self.employee_tab, text="Employees")
        
        self.ga01_weeks_tab = GA01WeeksTab(self.notebook)
        self.notebook.add(self.ga01_weeks_tab, text="GA01 Weeks")
        
        self.forecast_tab = ForecastTab(self.notebook)
        self.notebook.add(self.forecast_tab, text="Forecast")
        
        self.project_allocation_tab = ProjectAllocationTab(self.notebook)
        self.notebook.add(self.project_allocation_tab, text="Project Allocations")
        
        self.planned_changes_tab = PlannedChangesTab(self.notebook)
        self.notebook.add(self.planned_changes_tab, text="Planned Changes")
        
        self.forecast_visualization_tab = ForecastVisualization(self.notebook)
        self.notebook.add(self.forecast_visualization_tab, text="Visualizations")
        
        # Create menu
        self.create_menu()
        
        # Create status bar
        self.status_bar = ttk.Label(self, text="Ready", relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
    
    def init_settings(self):
        """Initialize settings with default values if not exists"""
        try:
            session = get_session()
            settings = session.query(Settings).first()
            if not settings:
                settings = Settings(fte_hours=34.5, contractor_hours=39.0)
                session.add(settings)
                session.commit()
            session.close()
        except Exception as e:
            print(f"Failed to initialize settings: {str(e)}")
    
    def create_menu(self):
        menubar = tk.Menu(self)
        self.config(menu=menubar)
        
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Settings", command=self.open_settings)
        file_menu.add_separator()
        file_menu.add_command(label="Import Employees", command=self.import_employees)
        file_menu.add_command(label="Import GA01 Weeks", command=self.import_ga01_weeks)
        file_menu.add_command(label="Export Forecast", command=self.export_forecast)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.quit)
        
        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="About", command=self.show_about)
    
    def import_employees(self):
        """Open dialog to import employees from CSV"""
        dialog = ImportEmployeesDialog(self)
    
    def import_ga01_weeks(self):
        messagebox.showinfo("Not Implemented", "Import GA01 Weeks feature coming soon!")
    
    def export_forecast(self):
        messagebox.showinfo("Not Implemented", "Export Forecast feature coming soon!")
    
    def open_settings(self):
        dialog = SettingsDialog(self)
        if dialog.get_settings():
            try:
                session = get_session()
                
                # Get or create settings
                settings = session.query(Settings).first()
                if not settings:
                    settings = Settings()
                    session.add(settings)
                
                # Update settings
                settings.fte_hours = dialog.get_settings()["fte_hours"]
                settings.contractor_hours = dialog.get_settings()["contractor_hours"]
                
                session.commit()
                session.close()
                
                messagebox.showinfo("Success", "Settings updated successfully.")
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update settings: {str(e)}")
    
    def show_about(self):
        messagebox.showinfo("About", "Forecast Tool v1.0\n\nA tool for managing employee forecasts.")

class SettingsDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.result = None
        
        self.title("Settings")
        self.geometry("400x200")
        self.resizable(True, True)
        
        # Make dialog modal
        self.transient(parent)
        self.grab_set()
        
        # Create form
        frame = ttk.Frame(self, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)
        
        # FTE Hours
        ttk.Label(frame, text="FTE Hours per Week:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.fte_hours_var = tk.DoubleVar(value=34.5)
        ttk.Entry(frame, textvariable=self.fte_hours_var, width=10).grid(row=0, column=1, sticky=tk.W, pady=5)
        
        # Contractor Hours
        ttk.Label(frame, text="Contractor Hours per Week:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.contractor_hours_var = tk.DoubleVar(value=39.0)
        ttk.Entry(frame, textvariable=self.contractor_hours_var, width=10).grid(row=1, column=1, sticky=tk.W, pady=5)
        
        # Buttons
        button_frame = ttk.Frame(frame)
        button_frame.grid(row=2, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="OK", command=self.on_ok, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=self.on_cancel, width=10).pack(side=tk.LEFT, padx=5)
        
        # Load current settings
        self.load_settings()
        
        # Center the dialog on the parent window
        self.center_on_parent()
    
    def center_on_parent(self):
        """Center the dialog on the parent window"""
        self.update_idletasks()
        
        # Get parent geometry
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        parent_x = self.parent.winfo_rootx()
        parent_y = self.parent.winfo_rooty()
        
        # Calculate position
        width = self.winfo_width()
        height = self.winfo_height()
        
        x = parent_x + (parent_width - width) // 2
        y = parent_y + (parent_height - height) // 2
        
        # Set position only (preserve size)
        self.geometry(f"+{x}+{y}")
    
    def load_settings(self):
        """Load settings from database"""
        try:
            session = get_session()
            settings = session.query(Settings).first()
            
            if settings:
                self.fte_hours_var.set(settings.fte_hours)
                self.contractor_hours_var.set(settings.contractor_hours)
            
            session.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load settings: {str(e)}")
    
    def on_ok(self):
        try:
            # Validate values
            fte_hours = float(self.fte_hours_var.get())
            contractor_hours = float(self.contractor_hours_var.get())
            
            if fte_hours <= 0:
                raise ValueError("FTE Hours must be greater than 0")
            if contractor_hours <= 0:
                raise ValueError("Contractor Hours must be greater than 0")
            
            self.result = {
                "fte_hours": fte_hours,
                "contractor_hours": contractor_hours
            }
            self.destroy()
        except ValueError as e:
            messagebox.showerror("Error", str(e))
    
    def on_cancel(self):
        self.result = None
        self.destroy()
    
    def get_settings(self):
        return self.result

class ImportEmployeesDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.filename = None
        self.data = None
        self.column_mappings = {}
        
        self.title("Import Employees")
        self.geometry("800x600")
        self.resizable(True, True)
        
        # Make dialog modal
        self.transient(parent)
        self.grab_set()
        
        # Create main frame
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # File selection
        file_frame = ttk.LabelFrame(main_frame, text="File Selection", padding="5")
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.file_label = ttk.Label(file_frame, text="No file selected")
        self.file_label.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        self.browse_button = ttk.Button(file_frame, text="Browse", command=self.browse_file)
        self.browse_button.pack(side=tk.RIGHT, padx=5)
        
        # Column mapping
        mapping_frame = ttk.LabelFrame(main_frame, text="Column Mapping", padding="5")
        mapping_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Required fields
        required_fields = ["name", "manager_code", "cost_center", "employment_type", "start_date"]
        self.mapping_vars = {}
        
        for i, field in enumerate(required_fields):
            ttk.Label(mapping_frame, text=f"{field.replace('_', ' ').title()}:").grid(row=i, column=0, sticky=tk.W, padx=5, pady=2)
            var = tk.StringVar()
            self.mapping_vars[field] = var
            combo = ttk.Combobox(mapping_frame, textvariable=var, state="readonly", width=30)
            combo.grid(row=i, column=1, sticky=tk.W, padx=5, pady=2)
        
        # Optional fields
        ttk.Label(mapping_frame, text="End Date:").grid(row=len(required_fields), column=0, sticky=tk.W, padx=5, pady=2)
        self.mapping_vars["end_date"] = tk.StringVar()
        combo = ttk.Combobox(mapping_frame, textvariable=self.mapping_vars["end_date"], state="readonly", width=30)
        combo.grid(row=len(required_fields), column=1, sticky=tk.W, padx=5, pady=2)
        
        # Preview
        preview_frame = ttk.LabelFrame(main_frame, text="Data Preview", padding="5")
        preview_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Create treeview for preview
        self.preview_tree = ttk.Treeview(preview_frame, show="headings")
        self.preview_tree.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
        
        # Add scrollbar to preview
        scrollbar = ttk.Scrollbar(preview_frame, orient="vertical", command=self.preview_tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.preview_tree.configure(yscrollcommand=scrollbar.set)
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Button(button_frame, text="Import", command=self.import_data, width=10).pack(side=tk.RIGHT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=self.destroy, width=10).pack(side=tk.RIGHT, padx=5)
        
        # Center the dialog
        self.center_on_parent()
    
    def center_on_parent(self):
        """Center the dialog on the parent window"""
        self.update_idletasks()
        parent_x = self.parent.winfo_rootx()
        parent_y = self.parent.winfo_rooty()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        
        width = self.winfo_width()
        height = self.winfo_height()
        
        x = parent_x + (parent_width - width) // 2
        y = parent_y + (parent_height - height) // 2
        
        self.geometry(f"+{x}+{y}")
    
    def browse_file(self):
        """Open file dialog to select CSV file"""
        try:
            filename = filedialog.askopenfilename(
                parent=self,
                title="Select CSV File",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )
            
            if filename:
                print(f"Selected file: {filename}")  # Debug log
                self.filename = filename
                self.file_label.config(text=os.path.basename(filename))
                self.load_preview()
        except Exception as e:
            print(f"Error in browse_file: {str(e)}")  # Debug log
            messagebox.showerror("Error", f"Failed to open file dialog: {str(e)}")
    
    def load_preview(self):
        """Load and display preview of CSV data"""
        try:
            print(f"Loading preview for file: {self.filename}")  # Debug log
            
            # Clear existing preview
            for item in self.preview_tree.get_children():
                self.preview_tree.delete(item)
            
            # Read CSV file
            with open(self.filename, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                headers = next(reader)  # Get headers
                self.data = list(reader)  # Get data rows
            
            print(f"Headers: {headers}")  # Debug log
            print(f"Number of data rows: {len(self.data)}")  # Debug log
            
            # Update column mappings
            for var in self.mapping_vars.values():
                var.set('')  # Clear previous selections
            
            # Update combobox values for all mapping fields
            for field, var in self.mapping_vars.items():
                combo_values = [''] + headers
                for widget in self.winfo_children()[0].winfo_children()[1].winfo_children():
                    if isinstance(widget, ttk.Combobox) and widget.cget('textvariable') == str(var):
                        widget['values'] = combo_values
            
            # Try to automatically map columns based on header names
            for field, var in self.mapping_vars.items():
                field_variations = [
                    field,
                    field.replace('_', ''),
                    field.replace('_', ' '),
                    field.title(),
                    field.replace('_', ' ').title(),
                ]
                
                for header in headers:
                    if header.lower().strip() in [v.lower() for v in field_variations]:
                        var.set(header)
                        break
            
            # Configure treeview columns
            self.preview_tree['columns'] = headers
            for col in headers:
                self.preview_tree.heading(col, text=col)
                self.preview_tree.column(col, width=100)
            
            # Add first 10 rows to preview
            for row in self.data[:10]:
                self.preview_tree.insert('', tk.END, values=row)
            
            print("Preview loaded successfully")  # Debug log
            
        except Exception as e:
            print(f"Error in load_preview: {str(e)}")  # Debug log
            messagebox.showerror("Error", f"Failed to load CSV file: {str(e)}")
    
    def import_data(self):
        """Import data from CSV file"""
        if not self.filename or not self.data:
            messagebox.showwarning("Warning", "Please select a CSV file first.")
            return
        
        # Validate mappings
        required_fields = ["name", "manager_code", "cost_center", "employment_type", "start_date"]
        missing_fields = []
        
        for field in required_fields:
            if not self.mapping_vars[field].get():
                missing_fields.append(field.replace('_', ' ').title())
        
        if missing_fields:
            messagebox.showwarning("Warning", f"Please map the following required fields: {', '.join(missing_fields)}")
            return
        
        try:
            # Get column indices for mapped fields
            mappings = {field: -1 for field in self.mapping_vars.keys()}
            with open(self.filename, 'r', encoding='utf-8-sig') as f:
                headers = next(csv.reader(f))
                for field, var in self.mapping_vars.items():
                    if var.get():
                        mappings[field] = headers.index(var.get())
            
            # Import data
            session = get_session()
            imported = 0
            errors = []
            
            for row in self.data:
                try:
                    # Parse dates
                    start_date = datetime.strptime(row[mappings["start_date"]], "%Y-%m-%d").date()
                    end_date = None
                    if mappings["end_date"] >= 0 and row[mappings["end_date"]].strip():
                        end_date = datetime.strptime(row[mappings["end_date"]], "%Y-%m-%d").date()
                    
                    # Create employee
                    employee = Employee(
                        name=row[mappings["name"]],
                        manager_code=row[mappings["manager_code"]],
                        cost_center=row[mappings["cost_center"]],
                        employment_type=row[mappings["employment_type"]],
                        start_date=start_date,
                        end_date=end_date
                    )
                    
                    session.add(employee)
                    imported += 1
                    
                except Exception as e:
                    errors.append(f"Row {imported + 1}: {str(e)}")
            
            session.commit()
            session.close()
            
            # Show results
            if errors:
                messagebox.showwarning("Import Complete", 
                    f"Imported {imported} employees with {len(errors)} errors.\n\n"
                    f"Errors:\n" + "\n".join(errors[:10]) +
                    ("\n..." if len(errors) > 10 else ""))
            else:
                messagebox.showinfo("Success", f"Successfully imported {imported} employees.")
            
            # Refresh employee list
            if isinstance(self.parent, ForecastApp):
                self.parent.employee_tab.load_employees()
            
            self.destroy()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to import data: {str(e)}")
            if 'session' in locals():
                session.close()

class AllocationDialog(tk.Toplevel):
    def __init__(self, parent, allocation=None):
        super().__init__(parent)
        self.parent = parent
        self.allocation = allocation
        self.result = None
        
        self.title("Add Allocation" if not allocation else "Edit Allocation")
        self.geometry("400x300")
        self.resizable(False, False)
        
        # Make dialog modal
        self.transient(parent)
        self.grab_set()
        
        # Create main frame
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Employee selection
        ttk.Label(main_frame, text="Employee:").pack(anchor=tk.W)
        self.employee_var = tk.StringVar()
        self.employee_combo = ttk.Combobox(main_frame, textvariable=self.employee_var, state="readonly")
        self.employee_combo['values'] = [emp.name for emp in self.parent.employees]
        self.employee_combo.pack(fill=tk.X, pady=(0, 10))
        
        # Project selection
        ttk.Label(main_frame, text="Project:").pack(anchor=tk.W)
        self.project_var = tk.StringVar()
        self.project_combo = ttk.Combobox(main_frame, textvariable=self.project_var, state="readonly")
        self.project_combo['values'] = [proj.name for proj in self.parent.projects]
        self.project_combo.pack(fill=tk.X, pady=(0, 10))
        
        # Hours
        ttk.Label(main_frame, text="Hours:").pack(anchor=tk.W)
        self.hours_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.hours_var).pack(fill=tk.X, pady=(0, 10))
        
        # Start date
        ttk.Label(main_frame, text="Start Date (YYYY-MM-DD):").pack(anchor=tk.W)
        self.start_date_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.start_date_var).pack(fill=tk.X, pady=(0, 10))
        
        # End date
        ttk.Label(main_frame, text="End Date (YYYY-MM-DD):").pack(anchor=tk.W)
        self.end_date_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.end_date_var).pack(fill=tk.X, pady=(0, 10))
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(button_frame, text="OK", command=self.on_ok, width=10).pack(side=tk.RIGHT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=self.destroy, width=10).pack(side=tk.RIGHT, padx=5)
        
        # If editing, populate fields
        if allocation:
            self.employee_var.set(allocation.employee.name)
            self.project_var.set(allocation.project.name)
            self.hours_var.set(str(allocation.hours))
            self.start_date_var.set(allocation.start_date.strftime("%Y-%m-%d"))
            self.end_date_var.set(allocation.end_date.strftime("%Y-%m-%d") if allocation.end_date else "")
        
        # Center the dialog
        self.center_on_parent()
    
    def on_ok(self):
        """Handle OK button click"""
        try:
            # Validate required fields
            if not self.employee_var.get():
                messagebox.showerror("Error", "Please select an employee")
                return
            if not self.project_var.get():
                messagebox.showerror("Error", "Please select a project")
                return
            if not self.hours_var.get():
                messagebox.showerror("Error", "Please enter hours")
                return
            if not self.start_date_var.get():
                messagebox.showerror("Error", "Please enter start date")
                return
            
            # Parse dates
            start_date = datetime.strptime(self.start_date_var.get(), "%Y-%m-%d").date()
            end_date = None
            if self.end_date_var.get():
                end_date = datetime.strptime(self.end_date_var.get(), "%Y-%m-%d").date()
            
            # Get employee and project
            employee = next(emp for emp in self.parent.employees if emp.name == self.employee_var.get())
            project = next(proj for proj in self.parent.projects if proj.name == self.project_var.get())
            
            # Create result
            self.result = {
                'employee_id': employee.id,
                'project_id': project.id,
                'hours': float(self.hours_var.get()),
                'start_date': start_date,
                'end_date': end_date
            }
            
            self.destroy()
            
        except ValueError as e:
            messagebox.showerror("Error", str(e))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create allocation: {str(e)}")
    
    def center_on_parent(self):
        """Center the dialog on its parent window"""
        self.update_idletasks()
        parent_x = self.parent.winfo_rootx()
        parent_y = self.parent.winfo_rooty()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        dialog_width = self.winfo_width()
        dialog_height = self.winfo_height()
        x = parent_x + (parent_width - dialog_width) // 2
        y = parent_y + (parent_height - dialog_height) // 2
        self.geometry(f"+{x}+{y}")

class AllocationsTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.allocations = []
        self.load_allocations()
        
        # Create toolbar
        toolbar = ttk.Frame(self)
        toolbar.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(toolbar, text="Add Allocation", command=self.add_allocation).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="Edit Allocation", command=self.edit_allocation).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="Delete Allocation", command=self.delete_allocation).pack(side=tk.LEFT, padx=2)
        
        # Create treeview
        self.tree = ttk.Treeview(self, columns=("employee", "project", "hours", "start_date", "end_date"), show="headings")
        self.tree.heading("employee", text="Employee")
        self.tree.heading("project", text="Project")
        self.tree.heading("hours", text="Hours")
        self.tree.heading("start_date", text="Start Date")
        self.tree.heading("end_date", text="End Date")
        
        # Configure column widths
        self.tree.column("employee", width=150)
        self.tree.column("project", width=150)
        self.tree.column("hours", width=100)
        self.tree.column("start_date", width=100)
        self.tree.column("end_date", width=100)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(self, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        # Pack widgets
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Load data
        self.refresh()
    
    def load_allocations(self):
        """Load allocations from database"""
        try:
            session = Session()
            self.allocations = session.query(Allocation).all()
            session.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load allocations: {str(e)}")
    
    def refresh(self):
        """Refresh the treeview with current data"""
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Add allocations
        for allocation in self.allocations:
            self.tree.insert("", tk.END, values=(
                allocation.employee.name,
                allocation.project.name,
                allocation.hours,
                allocation.start_date.strftime("%Y-%m-%d"),
                allocation.end_date.strftime("%Y-%m-%d") if allocation.end_date else ""
            ))
    
    def add_allocation(self):
        """Add a new allocation"""
        dialog = AllocationDialog(self)
        self.wait_window(dialog)
        
        if dialog.result:
            try:
                session = Session()
                allocation = Allocation(**dialog.result)
                session.add(allocation)
                session.commit()
                session.close()
                
                self.load_allocations()
                self.refresh()
                
                messagebox.showinfo("Success", "Allocation added successfully")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to add allocation: {str(e)}")
    
    def edit_allocation(self):
        """Edit selected allocation"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select an allocation to edit")
            return
        
        # Get selected allocation
        item = self.tree.item(selection[0])
        employee_name = item['values'][0]
        project_name = item['values'][1]
        
        allocation = next(
            (a for a in self.allocations 
             if a.employee.name == employee_name and a.project.name == project_name),
            None
        )
        
        if allocation:
            dialog = AllocationDialog(self, allocation)
            self.wait_window(dialog)
            
            if dialog.result:
                try:
                    session = Session()
                    for key, value in dialog.result.items():
                        setattr(allocation, key, value)
                    session.commit()
                    session.close()
                    
                    self.load_allocations()
                    self.refresh()
                    
                    messagebox.showinfo("Success", "Allocation updated successfully")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to update allocation: {str(e)}")
    
    def delete_allocation(self):
        """Delete selected allocation"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select an allocation to delete")
            return
        
        if messagebox.askyesno("Confirm", "Are you sure you want to delete this allocation?"):
            # Get selected allocation
            item = self.tree.item(selection[0])
            employee_name = item['values'][0]
            project_name = item['values'][1]
            
            allocation = next(
                (a for a in self.allocations 
                 if a.employee.name == employee_name and a.project.name == project_name),
                None
            )
            
            if allocation:
                try:
                    session = Session()
                    session.delete(allocation)
                    session.commit()
                    session.close()
                    
                    self.load_allocations()
                    self.refresh()
                    
                    messagebox.showinfo("Success", "Allocation deleted successfully")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to delete allocation: {str(e)}")

if __name__ == "__main__":
    app = ForecastApp()
    app.mainloop() 